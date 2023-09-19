# import openpyxl as xl;
  
# # opening the source excel file
# filename =r"C:\Users\astit\Git repos\open_excel\emp.xlsx"
# wb1 = xl.load_workbook(filename)
# ws1 = wb1.worksheets[0]
  
# # opening the destination excel file 
# filename1 =r"C:\Users\astit\Git repos\open_excel\new_emp.xlsx"
# wb2 = xl.load_workbook(filename1)
# ws2 = wb2.active
  
# # calculate total number of rows and 
# # columns in source excel file
# mr = ws1.max_row
# mc = ws1.max_column
  
# # copying the cell values from source 
# # excel file to destination excel file
# for i in range (1, mr + 1):
#     for j in range (1, mc + 1):
#         # reading cell value from source excel file
#         c = ws1.cell(row = i, column = j)
  
#         # writing the read value to destination excel file
#         ws2.cell(row = i, column = j).value = c.value
  
# # saving the destination excel file
# wb2.save(str(filename1))
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
from openpyxl import Workbook
xl1 = openpyxl.load_workbook('new_emp.xlsx')
# sheet you want to copy
# s = openpyxl.load_workbook(r'C:\Users\astit\Git repos\open_excel\new_emp.xlsx').active
# s._parent = xl1
# xl1._add_sheet(s)
# xl1.save(r'C:\Users\astit\Git repos\open_excel\new_emp.xlsx')
# print(xl1.sheetnames)
# xl1.copy_worksheet(xl1['Sheet1'])
# print(xl1.sheetnames)
# xl1.save('testing.xlsx')
ws = xl1.active
for cell in ws[1]:
    if cell.value == "Name":
        fullname_column = cell.column_letter
        break
for cell in ws[1]:
    if cell.value == "Salary":
        Salary_column = cell.column_letter
        break
firstname_column = ws.max_column + 1
lastname_column = ws.max_column + 2
new_salary_column=ws.max_column +3

ws.cell(row=1, column=firstname_column).value = "First Name"
ws.cell(row=1, column=lastname_column).value = "Last Name"
ws.cell(row=1, column=new_salary_column).value = "New_salary"


for cell in ws[fullname_column][1:]:
    fullname = cell.value.split(maxsplit=1)
    try:
        ws.cell(row=cell.row, column=firstname_column).value = fullname[0]
        ws.cell(row=cell.row, column=lastname_column).value = fullname[1]
    except IndexError:
        pass


for cell1 in ws[Salary_column][1:]:
    new_sal=cell1.value
    new_sal+=(new_sal/10)
    ws.cell(row=cell1.row,column=new_salary_column).value=new_sal

yellow = "00FFFF00"
ws.delete_cols(cell.column)
##ws.delete_cols(cell1.column)
for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
# ws['B1'].fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
x = ws.cell(row=1, column=1)
y = ws.cell(row=1, column=2)
z = ws.cell(row=1, column=3)
a=ws.cell(row=1,column=4)
b=ws.cell(row=1,column=5)
print(x.value)
mr = ws.max_row
mc = ws.max_column
for i in range (2, mr + 1):
    wb1=Workbook()

    s_name=i.__str__()
   ## wb1.create_sheet(s_name)
    ##new_s=wb1[s_name]
    new_s=wb1.active
    new_s.cell(row=1,column=1).value=x.value
    new_s.cell(row=1,column=1).fill=copy(x.fill)
    new_s.cell(row=1,column=2).value=y.value
    new_s.cell(row=1,column=3).value=z.value
    new_s.cell(row=1,column=4).value=a.value
    new_s.cell(row=1,column=5).value=b.value
    for j in range (1, mc + 1):
        c = ws.cell(row = i, column = j)
        new_s.cell(row = 2, column = j).value = c.value
    wb1.save(s_name+".xlsx") 
    
    

xl1.save('testing.xlsx')



